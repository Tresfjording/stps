from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from pathlib import Path

from blabla import build_tipper_dashboard_sheet


def test_build_tipper_dashboard_sheet_writes_player_progress(tmp_path) -> None:
    # Create a test workbook with sample data
    workbook_path = tmp_path / "demo.xlsx"
    wb = Workbook()

    ws = wb.create_sheet("f_ahh")
    ws.append(["Ukenr", "Dato", "hp", "up", "bp", "Totalt"])
    ws.append([1, datetime(2025, 8, 16), 130, 10, 20, 160])
    ws.append([2, datetime(2025, 8, 23), 200, 15, 25, 240])

    table = Table(displayName="t_ahh", ref="A1:F3")
    ws.add_table(table)

    wb.create_sheet("Dashboard")
    wb.save(workbook_path)

    # Change working directory to tmp_path so Dashboard.xlsm is created there
    import os
    original_cwd = os.getcwd()
    os.chdir(tmp_path)
    
    try:
        result_df = build_tipper_dashboard_sheet(str(workbook_path), sheet_name="Dashboard")

        assert not result_df.empty
        assert list(result_df.columns) == ["Tipper", "Dato", "hp", "up", "bp", "Totalt"]
        assert result_df.iloc[0]["Tipper"] == "AHH"
        assert result_df.iloc[0]["hp"] == 130
        assert result_df.iloc[1]["Totalt"] == 240

        # Verify data was written to Dashboard.xlsm (not the input file)
        dashboard_path = tmp_path / "Dashboard.xlsm"
        assert dashboard_path.exists(), "Dashboard.xlsm should be created"
        
        saved_wb = load_workbook(dashboard_path, data_only=True, read_only=False)
        saved_sheet = saved_wb["Dashboard"]
        assert saved_sheet.cell(2, 1).value == "AHH"
        assert saved_sheet.cell(2, 2).value == datetime(2025, 8, 16)
        assert saved_sheet.cell(3, 2).value == datetime(2025, 8, 23)
        saved_wb.close()
    finally:
        os.chdir(original_cwd)
