
print("JEG ER HER!!!")


import sqlite3
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

import os
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FixedLocator


def save_and_close_open_excel_workbooks(target_filename: str | None = None, save=True, close=True):
    """Try to save and close open Excel workbooks (Windows only).

    If `target_filename` is provided, only attempts to save/close the workbook
    whose `Name` matches (or contains) that filename. If not provided, acts on
    all open workbooks.

    Uses pywin32 (win32com). If no running Excel instance or pywin32 is
    unavailable, returns False and prints an informational message.
    """
    try:
        import win32com.client
        from win32com.client import GetActiveObject
    except Exception as exc:
        print(f"Excel automation not available (pywin32): {exc}")
        return False

    try:
        try:
            xl = GetActiveObject("Excel.Application")
        except Exception:
            return False

        count = xl.Workbooks.Count
        for i in range(count, 0, -1):
            try:
                wb = xl.Workbooks(i)
                name = getattr(wb, "Name", "<unknown>")
                if target_filename:
                    # match either exact or substring (handles full paths/opened copies)
                    if target_filename not in name:
                        continue

                if save:
                    wb.Save()
                    print(f"Saved workbook: {name}")
                if close:
                    wb.Close(SaveChanges=False)
                    print(f"Closed workbook: {name}")
            except Exception as e:
                print(f"Could not save/close workbook {i}: {e}")

        return True
    except Exception as e:
        print(f"Error interacting with Excel: {e}")
        return False


print("\nDEBUG INFO:")
print("Working dir:", os.getcwd())

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
print("BASE_DIR:", BASE_DIR)

db_path = os.path.join(BASE_DIR, "data", "stps.db")
schema_path = os.path.join(BASE_DIR, "database", "schema.sql")

print("DB path:", db_path)
print("Schema path:", schema_path)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

db_path = os.path.join(BASE_DIR, "data", "stps.db")
schema_path = os.path.join(BASE_DIR, "database", "schema.sql")

print("DB path:", db_path)
print("Schema path:", schema_path)

conn = sqlite3.connect(db_path)
cursor = conn.cursor()


# Rydd databasen før vi fyller på nytt
cursor.execute("DELETE FROM tips;")
cursor.execute("DELETE FROM matches;")
cursor.execute("DELETE FROM players;")
cursor.execute("DELETE FROM weeks;")
conn.commit()


with open(schema_path, "r") as f:
    sql_script = f.read()

#---cursor.executescript(sql_script)---

conn.commit()   # ✅ VIKTIG

print("\n--- SQL SCRIPT ---")
print(sql_script)

cursor.executescript(sql_script)
# --- Paths ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

db_path = os.path.join(BASE_DIR, "data", "stps.db")
schema_path = os.path.join(BASE_DIR, "database", "schema.sql")

# Sørg for at data-mappen finnes
os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)

# --- Database connection ---
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# --- Opprett tabeller ---
with open(schema_path, "r") as f:
    cursor.executescript(f.read())

conn.commit()

print("✅ Database opprettet!")


# --- RESET DATA (kun under utvikling!) ---




# --- LEGG INN UKE ---
cursor.execute(
    "INSERT OR IGNORE INTO weeks (week_number, date) VALUES (?, ?)",
    (1, "2026-08-01")
)

conn.commit()

print("Uke lagt inn!")

# --- LEGG INN UKE 2 ---
cursor.execute(
    "INSERT OR IGNORE INTO weeks (week_number, date) VALUES (?, ?)",
    (2, "2026-08-08")
)

conn.commit()

# --- FINN week_id ---
cursor.execute("SELECT id FROM weeks WHERE week_number = 2")
week_id = cursor.fetchone()[0]


# --- LEGG INN KAMPER ---
for i in range(1, 13):
    cursor.execute(
        "INSERT OR IGNORE INTO matches (week_id, match_number, result) VALUES (?, ?, ?)",
        (week_id, i, None)
    )

  # --- LEGG INN SPILLERE ---
players = ["AHH", "RB", "EB", "KAF", "ØG", "JH", "TOH", "UTG", "LEV"]

for p in players:
    cursor.execute("INSERT OR IGNORE INTO players (name) VALUES (?)", (p,))

conn.commit()

print("Spillere lagt inn!")


# --- LEGG INN TIPS ---
for name in players:
    cursor.execute("SELECT id FROM players WHERE name = ?", (name,))
    result = cursor.fetchone()

    if result is None:
        raise ValueError(f"Fant ikke spiller: {name}")

    player_id = result[0]

    # 🔥 BESTEM STRATEGI FØR SQL
    if name == "AHH":
        h, u, b = 50, 30, 20
    elif name == "RB":
        h, u, b = 70, 20, 10
    elif name == "EB":
        h, u, b = 30, 40, 30
    elif name == "KAF":
        h, u, b = 60, 20, 20
    elif name == "ØG":
        h, u, b = 40, 40, 20
    elif name == "JH":
        h, u, b = 55, 25, 20
    elif name == "TOH":
        h, u, b = 45, 35, 20
    elif name == "UTG":
        h, u, b = 50, 30, 20
    elif name == "LEV":
        h, u, b = 50, 30, 20
    else:
        h, u, b = 50, 30, 20

    # ✅ SQL KUN HER
    for match_number in range(1, 5):
        cursor.execute("""
            INSERT OR IGNORE INTO tips
            (player_id, week_id, match_number, h_percent, u_percent, b_percent)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (player_id, week_id, match_number, h, u, b))

conn.commit()

print("Tips lagt inn!")

print("Spillere lagt inn!")

print("Kamper lagt inn!")

# --- SETT RESULTATER ---
for i in range(1, 5):
    cursor.execute("""
        UPDATE matches
        SET result = 'H'
        WHERE week_id = ? AND match_number = ?
    """, (week_id, i))

conn.commit()

# --- TEST TIPS ---
cursor.execute("SELECT id FROM players WHERE name = ?", ("AHH",))
player_id = cursor.fetchone()[0]

# Legg inn tips for 4 kamper
# for match_number in range(1, 5):
#    cursor.execute("""
#        INSERT OR IGNORE INTO tips (player_id, week_id, match_number, h_percent, u_percent, b_percent)
#        VALUES (?, ?, ?, ?, ?, ?)
#    """, (player_id, week_id, match_number, 50, 30, 20))

conn.commit()

print("Test-tips lagt inn!")

# --- HENT RESULTAT ---
# Midlertidig test: sett kamp 1 = H
# Sett resultat på kamp 1–4


# --- SETT RESULTATER FØRST ---
for i in range(1, 5):
    cursor.execute("""

UPDATE matches
SET result = 'H'
WHERE week_id = 1 AND match_number = ?
    """, (i,))

conn.commit()


# --- SÅ henter du tips + resultater ---


cursor.execute("""
SELECT 
    t.player_id,
    t.h_percent,
    t.u_percent,
    t.b_percent,
    m.result
FROM tips t
JOIN matches m 
  ON t.week_id = m.week_id
 AND t.match_number = m.match_number
WHERE t.week_id = ?
""", (week_id,))




rows = cursor.fetchall()
print("ROWS (før beregning):", rows[:20])


totals = {}
correct = {}

for row in rows:
    print("DEBUG row:", row)   # 🔥 debug her

    player_id, h, u, b, result = row

    # --- POENG ---
    if result == "H":
        percent = h
        max_points = 65
    elif result == "U":
        percent = u
        max_points = 74
    else:
        percent = b
        max_points = 65

    points = min(percent, max_points)
    print(f"Player {player_id} fikk {points} poeng")

    # --- SUM ---
    totals[player_id] = totals.get(player_id, 0) + points

    # --- RETTE ---
    correct[player_id] = correct.get(player_id, 0)
    if result == "H" and h > 0:
        correct[player_id] += 1
    elif result == "U" and u > 0:
        correct[player_id] += 1
    elif result == "B" and b > 0:
        correct[player_id] += 1

for player_id, total in totals.items():
    correct_count = correct[player_id]

    bonus = 0
    if correct_count == 12:
        bonus = 1070
    elif correct_count == 11:
        bonus = 161
    elif correct_count == 10:
        bonus = 58

    cursor.execute("""
        INSERT OR IGNORE INTO weekly_results
        (player_id, week_id, total_points, correct, bonus)
        VALUES (?, ?, ?, ?, ?)
    """, (player_id, week_id, total, correct_count, bonus))

conn.commit()

history_data = []

for player_id in totals:
    history_data.append([
        player_id,
        totals[player_id],
        correct[player_id]
    ])

df_history = pd.DataFrame(history_data, columns=["PlayerID", "Poeng", "Rette"])

cursor.execute("SELECT id, name FROM players")
player_map = dict(cursor.fetchall())

df_history["Navn"] = df_history["PlayerID"].map(player_map)

# Rydd kolonner
df_history = df_history[["Navn", "Poeng", "Rette"]]

# --- DEBUG ---
print("\nDEBUG:")
print(f"Antall rader i rows: {len(rows)}")
print("Totals:", totals)
print("Correct:", correct)

# --- TOTAL ---
print("\nTOTAL POENG:")
for player_id, total in totals.items():
    print(f"Player {player_id}: {total} poeng")

# --- BONUS ---
print("\nBONUS:")
for player_id, count in correct.items():
    bonus = 0

    if count == 12:
        bonus = 1070

    

# --- LEADERBOARD ---
print("\nLEADERBOARD:")

cursor.execute("SELECT id, name FROM players")
player_names = dict(cursor.fetchall())

sorted_players = sorted(
    totals.items(),
    key=lambda x: (x[1], correct[x[0]]),
    reverse=True
)


for i, (player_id, total) in enumerate(sorted_players, start=1):
    name = player_names[player_id]
    print(f"{i}. {name}: {total} poeng")

winner_id, winner_points = sorted_players[0]
winner_name = player_names[winner_id]

print(f"\n🏆 Vinner uke {week_id}: {winner_name} med {winner_points} poeng!")



for row in cursor.fetchall():
    print(row)

cursor.execute("""
SELECT 
    p.name,
    SUM(r.total_points) as total,
    SUM(r.correct) as total_correct
FROM weekly_results r
JOIN players p ON r.player_id = p.id
GROUP BY p.name
ORDER BY total DESC, total_correct DESC
""")





for i, row in enumerate(cursor.fetchall(), start=1):
    name, total, correct_value = row
    print(f"{i}. {name}: {total} poeng ({correct_value} rette)")



rows = cursor.fetchall()
data = []

for player_id in totals:
    data.append([
        player_map[player_id],
        totals[player_id],
        correct[player_id]
    ])
    df_total = pd.DataFrame(data, columns=["Navn", "Poeng", "Rette"])

# hent data én gang
cursor.execute("""
SELECT 
    p.name,
    SUM(r.total_points) as total,
    SUM(r.correct) as total_correct
FROM weekly_results r
JOIN players p ON r.player_id = p.id
GROUP BY p.name
ORDER BY total DESC, total_correct DESC
""")
rows = cursor.fetchall()

print("\nSAMMENLAGT:")



for i, (name, total, correct) in enumerate(rows, start=1):
    print(f"{i}. {name}: {total} poeng ({correct} rette)")

import pandas as pd


def load_stps_tolk_data(filepath="stps_tolk.xlsx"):
    try:
        df_hovedtabell = pd.read_excel(filepath, sheet_name="Hovedtabell", engine="openpyxl")
    except Exception as exc:
        print(f"Warning: could not load '{filepath}' Hovedtabell: {exc}")
        df_hovedtabell = pd.DataFrame()

    try:
        df_kuponger_raw = pd.read_excel(filepath, sheet_name="kuponger", engine="openpyxl")
    except Exception as exc:
        print(f"Warning: could not load '{filepath}' kuponger: {exc}")
        df_kuponger_raw = pd.DataFrame()

    return df_hovedtabell, df_kuponger_raw


def sanitize_sheet_name(name: str) -> str:
    invalid = "\\/*[]:?"
    cleaned = "".join("_" if ch in invalid else ch for ch in str(name))
    return cleaned[:31]


def add_numeric_column_charts(writer, df_hovedtabell):
    if df_hovedtabell.empty or "Navn" not in df_hovedtabell.columns:
        return

    ws = writer.sheets.get("Hovedtabell_STPS")
    if ws is None:
        return

    numeric_cols = [
        c for c in df_hovedtabell.columns
        if c != "Navn" and pd.api.types.is_numeric_dtype(df_hovedtabell[c])
    ]

    chart_row = 1
    for col_name in numeric_cols:
        col_idx = df_hovedtabell.columns.get_loc(col_name) + 1
        chart = BarChart()
        chart.title = f"{col_name}"
        chart.y_axis.title = col_name
        chart.x_axis.title = "Spiller"
        chart.height = 10
        chart.width = 18

        max_row = len(df_hovedtabell) + 1
        chart.add_data(Reference(ws, min_col=col_idx, min_row=1, max_row=max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=max_row))
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        ws.add_chart(chart, f"K{chart_row}")
        chart_row += 18


def build_sammenlagt_chart_pages(df_sammen):
    if df_sammen.empty or "Navn" not in df_sammen.columns:
        return []

    top_rows = df_sammen.head(15)
    numeric_cols = [
        c for c in top_rows.columns
        if c != "Navn" and pd.api.types.is_numeric_dtype(top_rows[c])
    ]
    if not numeric_cols:
        return []

    pages = []
    for _, row in top_rows.iterrows():
        player = str(row["Navn"])
        values = [float(row[c]) if pd.notna(row[c]) else 0.0 for c in numeric_cols]
        pages.append((player, numeric_cols, values))

    preferred_order = ["AHH", "RB", "EB", "KAF", "ØG", "JH", "TOH", "UTG", "LEV"]
    order_map = {name: idx for idx, name in enumerate(preferred_order)}
    pages.sort(key=lambda page: order_map.get(page[0], len(preferred_order)))
    return pages


def show_stps_charts_window(df_sammen):
    pages = build_sammenlagt_chart_pages(df_sammen)
    if not pages:
        return

    root = tk.Tk()
    root.title("Sammenlagt diagrammer")
    root.geometry("1200x760")

    title_var = tk.StringVar()
    title_label = ttk.Label(root, textvariable=title_var, font=("Segoe UI", 14, "bold"))
    title_label.pack(fill="x", padx=10, pady=6)

    fig = plt.Figure(figsize=(12, 7), dpi=100)
    ax = fig.add_subplot(111)
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.get_tk_widget().pack(expand=True, fill="both")

    instr = ttk.Label(root, text="Bruk piltastene ← og → for å bla gjennom spillere.")
    instr.pack(fill="x", padx=10, pady=4)

    current_index = {"value": 0}

    def draw_page(index):
        index %= len(pages)
        current_index["value"] = index
        player, metrics, values = pages[index]
        ax.clear()

        ax.set_title(f"{player} - Sammenlagt", pad=18)
        bars = ax.bar(metrics, values, color=plt.cm.Blues([0.3 + 0.7 * i / max(1, len(values) - 1) for i in range(len(values))]))
        ax.set_xlabel("")
        ax.set_ylabel("")
        ax.set_xticks(range(len(metrics)))
        ax.xaxis.set_major_locator(FixedLocator(range(len(metrics))))
        ax.set_xticklabels(metrics, rotation=30, ha="right")
        ax.margins(x=0.02)
        ax.set_ylim(0, max(values) * 1.12 if values else 1)

        for bar in bars:
            height = bar.get_height()
            ax.annotate(
                f"{height:.0f}",
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 5),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=10,
                fontweight="bold"
            )

        title_var.set(f"{player} ({index + 1}/{len(pages)})")
        canvas.draw()

    def on_key(event):
        if event.keysym == "Right":
            draw_page(current_index["value"] + 1)
        elif event.keysym == "Left":
            draw_page(current_index["value"] - 1)

    root.bind("<Left>", on_key)
    root.bind("<Right>", on_key)

    draw_page(0)
    root.mainloop()


def add_player_stps_charts(writer, df_hovedtabell):
    if df_hovedtabell.empty or "Navn" not in df_hovedtabell.columns:
        return

    numeric_cols = [
        c for c in df_hovedtabell.columns
        if c != "Navn" and pd.api.types.is_numeric_dtype(df_hovedtabell[c])
    ]
    if not numeric_cols:
        return

    for _, player_row in df_hovedtabell.iterrows():
        player = player_row["Navn"]
        sheet_name = sanitize_sheet_name(f"{player}_STPS")
        player_df = pd.DataFrame({
            "Metrikk": numeric_cols,
            "Verdi": [player_row[c] for c in numeric_cols]
        })

        player_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        if ws is None:
            continue

        chart = BarChart()
        chart.title = f"{player} - STPS"
        chart.y_axis.title = "Verdi"
        chart.x_axis.title = "Metrikk"
        chart.height = 12
        chart.width = 24

        max_row = len(player_df) + 1
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=max_row))
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        ws.add_chart(chart, "D2")


def add_sammenlagt_top_chart(writer, df_sammen_sorted, chart_col):
    ws = writer.sheets.get("Sammenlagt")
    if ws is None or chart_col not in df_sammen_sorted.columns:
        return

    max_rows = min(len(df_sammen_sorted), 15)
    data = Reference(ws, min_col=df_sammen_sorted.columns.get_loc(chart_col) + 1, min_row=1, max_row=max_rows + 1)
    categories = Reference(ws, min_col=1, min_row=2, max_row=max_rows + 1)

    chart = BarChart()
    chart.title = f"Sammenlagt: {chart_col} for rad 2–{max_rows + 1}"
    chart.y_axis.title = chart_col
    chart.x_axis.title = "Tipper"
    chart.height = 16
    chart.width = 28

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    if chart.series:
        colors = [
            "FFB2DD", "FFCCE5", "FFD6F0", "E2B5FF", "C2A3FF",
            "A39BFF", "7B8BFF", "5D7BFF", "3C6CFF", "1F5FFF",
            "0A4CDA", "0042B3", "003A8C", "003366", "002B4F"
        ]
        series = chart.series[0]
        for idx in range(max_rows):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = colors[idx % len(colors)]
            series.dPt.append(dp)

    ws.add_chart(chart, "E2")


# Hent tall fra stps_tolk.xlsx og merge dem inn før vi skriver Sammenlagt
(df_hovedtabell, df_kuponger_raw) = load_stps_tolk_data()

df = pd.DataFrame(rows, columns=["Navn", "Poeng", "Rette"])

if not df_hovedtabell.empty:
    stps_columns = [
        "Navn",
        "hp",
        "up",
        "bp",
        "straff",
        "tp",
        "Rank",
        "Antall rette",
        "Bonus",
        "Summert",
        "Antall 12",
        "Antall 11",
        "Antall 10",
        "10-premie",
        "11-premie",
        "12-premie",
        "Premie",
        "Totalt"
    ]
    available = [c for c in stps_columns if c in df_hovedtabell.columns]
    if len(available) > 1:
        df = df.merge(df_hovedtabell[available], on="Navn", how="left")
        print("Merged stps_tolk Hovedtabell columns into Sammenlagt:", [c for c in available if c != "Navn"])
    else:
        print("Warning: No matching Hovedtabell columns available for merge.")

try:
    df.to_excel("tippelag.xlsx", sheet_name="Sammenlagt", index=False)
except PermissionError:
    df.to_excel("tippelag_fallback.xlsx", sheet_name="Sammenlagt", index=False)
    print("Could not write to tippelag.xlsx (file open). Wrote to tippelag_fallback.xlsx instead.")

unique_rows = set(rows)
print("Unike rader:", len(unique_rows))

# Sorter Sammenlagt etter poeng
if "Poeng" in df.columns:
    df_total = df.sort_values(by="Poeng", ascending=False)
else:
    df_total = df

# --- HISTORIKK ---
import pandas as pd

print("Starter Excel...")

# Beregn per-uke poengfordeling per tipskategori (hp, hu, hb) og totalt
# Vi summerer poeng per kamp for hver tipper basert på resultatet i matches
cursor.execute("""
    SELECT p.name, w.week_number, t.h_percent, t.u_percent, t.b_percent, m.result
    FROM tips t
    JOIN players p ON t.player_id = p.id
    JOIN matches m ON t.week_id = m.week_id AND t.match_number = m.match_number
    JOIN weeks w ON t.week_id = w.id
    ORDER BY p.name, w.week_number
""")

tip_rows = cursor.fetchall()

weekly = {}
for name, week, h, u, b, result in tip_rows:
    key = (name, week)
    if key not in weekly:
        weekly[key] = {"hp": 0, "hu": 0, "hb": 0, "Totalt": 0}

    if result is None:
        continue

    if result == "H":
        pts = min((h or 0), 65)
        weekly[key]["hp"] += pts
    elif result == "U":
        pts = min((u or 0), 74)
        weekly[key]["hu"] += pts
    else:
        pts = min((b or 0), 65)
        weekly[key]["hb"] += pts

    weekly[key]["Totalt"] += pts

# Bygg DataFrame for Historikk
hist_rows = []
for (name, week), vals in sorted(weekly.items(), key=lambda x: (x[0][0], x[0][1])):
    hist_rows.append([name, week, vals["hp"], vals["hu"], vals["hb"], vals["Totalt"]])

df_hist = pd.DataFrame(hist_rows, columns=["Navn", "Uke", "hp", "hu", "hb", "Totalt"]) 

# Hent tall fra stps_tolk.xlsx
(df_hovedtabell, df_kuponger_raw) = load_stps_tolk_data()

# ✅ START WRITER FØRST
out_file = "tippelag.xlsx"
# If Excel is running and `tippelag.xlsx` is open, save & close that workbook only.
try:
    if os.name == "nt":
        save_and_close_open_excel_workbooks(target_filename=out_file, save=True, close=True)
except Exception as exc:
    print(f"Could not save/close target Excel workbook {out_file}: {exc}")
try:
    writer = pd.ExcelWriter(out_file, engine="openpyxl")
except PermissionError:
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = f"tippelag_fallback_{timestamp}.xlsx"
    writer = pd.ExcelWriter(out_file, engine="openpyxl")
    print(f"Could not open tippelag.xlsx (file may be open). Using {out_file} instead.")

# ✅ Sammenlagt
# Lag en versjon av Sammenlagt hvor området A1:T16 (header + 15 rader)
# er sortert etter kolonne T (bruker 'Totalt' hvis tilgjengelig, ellers 'Poeng').
df_sammen = df_total.copy()

# Sørg for minst 20 kolonner (A..T)
while df_sammen.shape[1] < 20:
    df_sammen[f"Extra_{df_sammen.shape[1]+1}"] = ""

if "Totalt" in df_sammen.columns:
    sort_col = "Totalt"
elif "Poeng" in df_sammen.columns:
    sort_col = "Poeng"
else:
    sort_col = df_sammen.columns[1]

# Sorter topp 15 rader etter sort_col, behold resten i opprinnelig rekkefølge
top_n = 15
top = df_sammen.head(top_n).sort_values(by=sort_col, ascending=False)
rest = df_sammen.iloc[top_n:]
df_sammen_sorted = pd.concat([top, rest], ignore_index=True)

# Begrens til kolonnene A..T (første 20 kolonner) ved skriving
cols_to_write = df_sammen_sorted.columns[:20]
df_sammen_sorted[cols_to_write].to_excel(writer, sheet_name="Sammenlagt", index=False)
add_sammenlagt_top_chart(writer, df_sammen_sorted, sort_col)


# ✅ Historikk
df_hist.to_excel(writer, sheet_name="Historikk", index=False)

# ✅ HENT STPS-TALL
if not df_hovedtabell.empty:
    df_hovedtabell.to_excel(writer, sheet_name="Hovedtabell_STPS", index=False)
    add_numeric_column_charts(writer, df_hovedtabell)
    add_player_stps_charts(writer, df_hovedtabell)

if not df_kuponger_raw.empty:
    df_kuponger_raw.to_excel(writer, sheet_name="Kuponger_STPS", index=False)

# ✅ KUPONG-ARK (input)
players = df_total["Navn"].tolist()
kamp_antall = 12

# Forsøk å hente kampdatoer fra det importerte 'kuponger' arket
# og bruk dem som radetiketter (første kolonne) i det transponerte arket.
date_headers = []
if not df_kuponger_raw.empty:
    raw_wb = load_workbook("stps_tolk.xlsx", data_only=True)
    raw_ws = raw_wb["kuponger"] if "kuponger" in raw_wb.sheetnames else raw_wb.active
    for row in raw_ws.iter_rows(min_row=1, max_row=50, values_only=True):
        for val in row:
            if val is None:
                continue
            if isinstance(val, (datetime.date, datetime.datetime)):
                date_headers.append(val.date().isoformat() if isinstance(val, datetime.datetime) else val.isoformat())
            else:
                try:
                    maybe_date = pd.to_datetime(val, errors="coerce")
                    if not pd.isna(maybe_date):
                        date_headers.append(maybe_date.date().isoformat())
                except Exception:
                    pass
    date_headers = [d for d in dict.fromkeys(date_headers) if d is not None and int(d[:4]) >= 2000]

if len(date_headers) >= kamp_antall:
    date_headers = date_headers[:kamp_antall]
elif len(date_headers) >= 1:
    date_headers = [date_headers[0]] * kamp_antall
else:
    date_headers = [f"K{i}" for i in range(1, kamp_antall + 1)]

# Bygg et transponert kupongark: Dato/label i første kolonne, spillere som kolonner.
kupong_df = pd.DataFrame(index=date_headers[:kamp_antall], columns=players + ["Sum"])

# Fyll med tomme verdier.
kupong_df[:] = ""

kupong_df.index.name = "Dato"

df_kupong = kupong_df.reset_index()

df_kupong.to_excel(writer, sheet_name="Kuponger", index=False)

# ✅ Spillere
for player in df_hist["Navn"].dropna().unique():
    sheet_name = str(player).replace("/", "").replace("\\", "").replace(":", "")[:31]
    df_player = df_hist[df_hist["Navn"] == player]

    if not df_player.empty:
        df_player.to_excel(writer, sheet_name=sheet_name, index=False)

# ✅ LAGRE PÅ SLUTT
writer.close()

try:
    show_stps_charts_window(df_sammen_sorted)
except Exception as exc:
    print(f"Kunne ikke åpne diagramvindu: {exc}")

print("✅ Excel HELT OK!")
