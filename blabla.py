import pandas as pd
import matplotlib.pyplot as plt
import io
import base64
import sys
import os
import shutil
import subprocess
import importlib.util
import logging
import json
import html
from pathlib import Path
from string import Template
import warnings
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from matplotlib.patches import Patch

warnings.filterwarnings(
    "ignore",
    message=".*extension is not supported and will be removed.*",
    category=UserWarning,
)
warnings.filterwarnings(
    "ignore",
    message=".*wmf image format is not supported so the image is being dropped.*",
    category=UserWarning,
)


logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("stps_report")

WORKBOOK_FILE = "stps_tolk.xlsx"
SHEET_NAME = "Hovedtabell"
REPORT_TEMPLATE_PATH = Path("data") / "report_template.html"
PLAYER_HISTORY_FILE = Path("data") / "player_history.json"
LOGO_FILENAME = "st_logo_hvit_bgr.png"
LOGO_FILENAME_PATH = Path("data") / LOGO_FILENAME
LOGO_ENV_VAR = "STPS_LOGO_PATH"
SOURCE_CANDIDATES = [
    "stps_2026_RTM.xlsm",
    "STPS_2026_RTM.xlsm",
    "stps_2027_RTM.xlsm",
    "STPS_2027_RTM.xlsm",
    "stps_2027.xlsm",
    "STPS_2027.xlsm",
    "STPS 2027.xlsm",
    "stps_tolk.xlsx.xlsm",
]
WEEKLY_CANDIDATES = [
    "Sist_uke",
    "Alle_Siste_Ukes_Poeng",
    "Ukevinnere",
    "UkasVinner",
    "ukevinneretabell",
]

file = WORKBOOK_FILE
sheet_name = SHEET_NAME
refresh_on_start = "--refresh" in sys.argv


def resolve_workbook_path(preferred_file: str) -> str:
    preferred_path = Path(preferred_file)
    if preferred_path.exists():
        return str(preferred_path)

    xlsm_files = sorted(
        p for p in Path(".").glob("*.xlsm") if not p.name.startswith("~$")
    )
    if not xlsm_files:
        raise FileNotFoundError("Fant ingen .xlsm-filer i arbeidsmappen.")

    # Fuzzy match filename regardless of case, spaces, underscores, etc.
    def normalize_name(name: str) -> str:
        return re.sub(r"[^a-z0-9]", "", name.lower())

    preferred_norm = normalize_name(preferred_path.stem)
    exact_like = [p for p in xlsm_files if normalize_name(p.stem) == preferred_norm]
    if exact_like:
        return str(exact_like[0])

    # Prefer files that look like STPS annual workbooks.
    stps_candidates = [p for p in xlsm_files if "stps" in p.name.lower()]
    selected = stps_candidates[0] if stps_candidates else xlsm_files[0]

    logger.warning(f"Advarsel: '{preferred_file}' ble ikke funnet. Bruker '{selected.name}' i stedet.")
    return str(selected)


def update_stps_tolk_on_start(target_file: str, target_sheet: str) -> pd.DataFrame | None:
    source_file = None
    for candidate in SOURCE_CANDIDATES:
        if Path(candidate).exists():
            source_file = candidate
            break

    if source_file is None:
        logger.info("Info: Fant ingen 2027-kilde for oppdatering. Bruker eksisterende stps_tolk.xlsx.")
        return None

    try:
        source_df = load_named_range_to_df(source_file, "Sammendrag")

        # Write plain values (no formatting) so stps_tolk.xlsx stays a simple data store.
        write_df = source_df.copy().where(pd.notna(source_df), "")

        with pd.ExcelWriter(target_file, engine="openpyxl", mode="w") as writer:
            write_df.to_excel(writer, sheet_name=target_sheet, index=False)
        logger.info(f"Oppdaterte '{target_file}' fra '{source_file}' (område: Sammendrag).")
        return source_df
    except PermissionError:
        logger.warning(f"Advarsel: '{target_file}' er låst og kunne ikke skrives.")
        logger.info("Bruker oppdaterte data i minnet for denne kjøringen.")
        return source_df
    except Exception as exc:
        logger.warning(f"Advarsel: Kunne ikke oppdatere '{target_file}' fra '{source_file}': {exc}")
        logger.info("Bruker eksisterende stps_tolk.xlsx videre i kjøringen.")
        return None


def find_source_workbook() -> str | None:
    for candidate in SOURCE_CANDIDATES:
        if Path(candidate).exists():
            return candidate

    return None


def refresh_linked_workbooks_via_excel(source_file: str, target_file: str) -> bool:
    """Refresh linked Excel data by opening source first, then target, through Excel COM."""
    if os.name != "nt":
        return False

    try:
        import pythoncom  # type: ignore
        import win32com.client as win32  # type: ignore
    except Exception as exc:
        logger.info(f"Info: Excel COM ikke tilgjengelig for link-oppdatering: {exc}")
        return False

    source_path = str(Path(source_file).resolve())
    target_path = str(Path(target_file).resolve())

    excel = None
    source_wb = None
    target_wb = None

    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        # Open source first so external references in target can resolve against an open workbook.
        source_wb = excel.Workbooks.Open(source_path, UpdateLinks=3, ReadOnly=True)
        source_wb.RefreshAll()
        excel.CalculateFullRebuild()

        target_wb = excel.Workbooks.Open(target_path, UpdateLinks=3, ReadOnly=False)
        target_wb.RefreshAll()
        excel.CalculateFullRebuild()
        target_wb.Save()

        logger.info("Info: Oppdaterte linker i stps_tolk.xlsx via Excel.")
        return True
    except Exception as exc:
        logger.warning(f"Advarsel: Klarte ikke å oppdatere linker via Excel: {exc}")
        return False
    finally:
        if target_wb is not None:
            try:
                target_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if source_wb is not None:
            try:
                source_wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()  # type: ignore[name-defined]
        except Exception:
            pass


def build_weekly_filename(
    base_name: str,
    extension: str,
    report_week: int,
    report_year: int,
) -> str:
    # Avoid accidental double separators when callers pass names like "report_".
    clean_base = re.sub(r"[_\-\s]+$", "", base_name)
    if not clean_base:
        clean_base = "report"
    return f"{clean_base}_uke_{report_week}_{report_year}{extension}"


def create_weekly_workbook_copy(
    file_path: str,
    report_week: int,
    report_year: int,
) -> str | None:
    source_path = Path(file_path)
    if not source_path.exists():
        return None

    copy_name = build_weekly_filename(
        source_path.stem,
        source_path.suffix,
        report_week,
        report_year,
    )
    copy_path = source_path.with_name(copy_name)
    shutil.copy2(source_path, copy_path)
    return str(copy_path)


def render_report_html(
    template_path: Path,
    logo_src: str,
    report_week: int,
    report_pdf_filename: str,
    pdf_ready_hint: bool,
    report_updated_at: str,
    summary_section_html: str,
    weekly_section_html: str,
    player_history_section_html: str,
) -> str:
    template = Template(template_path.read_text(encoding="utf-8"))
    return template.substitute(
        logo_src=logo_src,
        report_week=report_week,
        report_pdf_filename=report_pdf_filename,
        pdf_status_text=(
            "PDF-eksport klar" if pdf_ready_hint else "PDF krever weasyprint eller wkhtmltopdf"
        ),
        report_updated_at=report_updated_at,
        summary_section_html=summary_section_html,
        weekly_section_html=weekly_section_html,
        player_history_section_html=player_history_section_html,
    )


def resolve_logo_path() -> Path:
    """Resolve logo path from env var and common project locations."""
    candidates: list[Path] = []

    env_logo_path = os.environ.get(LOGO_ENV_VAR)
    if env_logo_path:
        candidates.append(Path(env_logo_path).expanduser())

    base_dir = Path(__file__).resolve().parent
    cwd = Path.cwd()
    candidates.extend(
        [
            base_dir / LOGO_FILENAME,
            base_dir / "data" / LOGO_FILENAME,
            cwd / LOGO_FILENAME,
            cwd / "data" / LOGO_FILENAME,
        ]
    )

    for candidate in candidates:
        if candidate.is_file():
            return candidate

    return base_dir / LOGO_FILENAME


def build_logo_src(logo_path: Path) -> str:
    """Return a data URI for the logo to avoid filesystem encoding issues in PDF renderers."""
    if logo_path.is_file():
        try:
            encoded_logo = base64.b64encode(logo_path.read_bytes()).decode("ascii")
            return f"data:image/png;base64,{encoded_logo}"
        except Exception as exc:
            logger.warning(f"Advarsel: Klarte ikke å base64-enkode logo '{logo_path}': {exc}")

    logger.warning(
        f"Advarsel: Fant ikke logo '{logo_path}'. "
        f"Sett {LOGO_ENV_VAR} til riktig filsti hvis logoen ligger et annet sted. "
        "Fortsetter uten logobilde."
    )
    empty_svg = "<svg xmlns='http://www.w3.org/2000/svg' width='1' height='1'></svg>"
    encoded_svg = base64.b64encode(empty_svg.encode("utf-8")).decode("ascii")
    return f"data:image/svg+xml;base64,{encoded_svg}"


def table_to_df(
    wb,
    table_name: str | None = None,
    allow_any_totalt_table: bool = True,
) -> pd.DataFrame | None:
    for ws in wb.worksheets:
        tables = ws.tables
        if not tables:
            continue

        if table_name:
            for name, table in tables.items():
                if name.lower() == table_name.lower():
                    table_ref = table.ref if hasattr(table, "ref") else str(table)
                    values = [[cell.value for cell in row] for row in ws[table_ref]]
                    if values:
                        header, *data = values
                        return pd.DataFrame(data, columns=header)

        if allow_any_totalt_table:
            for table in tables.values():
                table_ref = table.ref if hasattr(table, "ref") else str(table)
                values = [[cell.value for cell in row] for row in ws[table_ref]]
                if not values:
                    continue
                header, *data = values
                if "Totalt" in [str(h) if h is not None else "" for h in header]:
                    return pd.DataFrame(data, columns=header)

    return None


def load_table_from_workbook(file_path: str, table_name: str) -> pd.DataFrame:
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    try:
        table_df = table_to_df(wb, table_name=table_name, allow_any_totalt_table=False)
        if table_df is None:
            raise ValueError(f"Fant ikke tabell '{table_name}' i {file_path}.")
        return table_df
    finally:
        wb.close()


def load_named_range_to_df(
    file_path: str,
    range_name: str,
    allow_generic_totalt_fallback: bool = True,
) -> pd.DataFrame:
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)

    def resolve_defined_name(name: str) -> str | None:
        all_names = list(wb.defined_names.keys())

        if name in wb.defined_names:
            return name

        lowered = {n.lower(): n for n in all_names}
        if name.lower() in lowered:
            return lowered[name.lower()]

        # Common case: user writes "Sammendrag", workbook contains "sammendraget".
        for n in all_names:
            if n.lower().startswith(name.lower()) or name.lower().startswith(n.lower()):
                return n

        # Flexible match for names with separators/extra words, e.g. "Sist_uke" -> "Alle_Siste_Ukes_Poeng".
        def normalize_token(text: str) -> str:
            return re.sub(r"[^a-z0-9]", "", text.lower())

        target = normalize_token(name)
        for n in all_names:
            candidate = normalize_token(n)
            if target and target in candidate:
                return n

        # Token match: "Sist_uke" should match names like "Alle_Siste_Ukes_Poeng".
        query_tokens = [t for t in re.split(r"[^a-z0-9]+", name.lower()) if t]
        for n in all_names:
            candidate_tokens = [t for t in re.split(r"[^a-z0-9]+", n.lower()) if t]
            if query_tokens and all(
                any(ct.startswith(qt) or qt.startswith(ct) for ct in candidate_tokens)
                for qt in query_tokens
            ):
                return n

        return None

    resolved_name = resolve_defined_name(range_name)
    if resolved_name is None:
        wb.close()
        raise ValueError(
            f"Fant ikke navngitt område '{range_name}' i {file_path}. "
            f"Tilgjengelige navn: {', '.join(wb.defined_names.keys())}"
        )

    defn = wb.defined_names[resolved_name]
    rows = []

    for sheet_name, cell_range in defn.destinations:
        ws = wb[sheet_name]
        for row in ws[cell_range]:
            rows.append([cell.value for cell in row])

    if not rows:
        # Some workbooks keep stale defined names (#REF!). Fall back to tables.
        fallback_df = table_to_df(
            wb,
            table_name=range_name,
            allow_any_totalt_table=allow_generic_totalt_fallback,
        )
        wb.close()
        if fallback_df is not None:
            return fallback_df
        raise ValueError(f"Navngitt område '{range_name}' er tomt i {file_path}.")

    header, *data = rows
    df = pd.DataFrame(data, columns=header)
    wb.close()
    return df


def load_weekly_winner_counts_from_named_range(
    file_path: str,
    range_name: str = "Ukevinnere",
) -> pd.DataFrame | None:
    """Build a fallback weekly scoreboard from a winner history range.

    Expected shape is date in first column and winner code/name in second column.
    The function aggregates winner occurrences into columns Navn/Totalt so it can
    reuse the regular section renderer.
    """
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    try:
        if range_name not in wb.defined_names:
            lowered = {n.lower(): n for n in wb.defined_names.keys()}
            resolved = lowered.get(range_name.lower())
        else:
            resolved = range_name

        if not resolved:
            return None

        defn = wb.defined_names[resolved]
        rows: list[list[object]] = []
        for sheet_name, cell_range in defn.destinations:
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                rows.append([cell.value for cell in row])

        if not rows:
            return None

        winners: list[str] = []
        for row in rows:
            if len(row) < 2:
                continue
            winner = row[1]
            if winner is None:
                continue

            winner_text = str(winner).strip()
            if not winner_text or winner_text.lower() in {"nan", "none", "null"}:
                continue

            winners.append(winner_text.upper())

        if not winners:
            return None

        counts = pd.Series(winners, dtype="object").value_counts().reset_index()
        counts.columns = ["Navn", "Totalt"]
        return counts
    finally:
        wb.close()


def load_latest_weekly_from_player_tables(file_path: str) -> pd.DataFrame | None:
    """Build a weekly table from the latest row in each per-player history table.

    Expected source tables are in sheets like f_rb/f_og and have columns such as
    Ukenr, Dato, hp, up, bp, straff, tp, Rank, Endring.
    """
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    try:
        rows_out: list[dict[str, object]] = []

        for ws in wb.worksheets:
            if not ws.tables:
                continue

            for table_name, table in ws.tables.items():
                table_lower = table_name.lower()
                if not table_lower.startswith("t_"):
                    continue

                table_ref = table.ref if hasattr(table, "ref") else str(table)
                values = [[cell.value for cell in row] for row in ws[table_ref]]
                if not values:
                    continue

                header, *data = values
                if not data:
                    continue

                df_local = pd.DataFrame(data, columns=header)
                df_local.columns = [
                    str(c).replace("\xa0", " ").replace("\n", " ").strip() if c is not None else ""
                    for c in df_local.columns
                ]

                if "tp" not in df_local.columns or "Ukenr" not in df_local.columns:
                    continue

                df_local["tp"] = pd.to_numeric(df_local["tp"], errors="coerce")
                df_local["Ukenr"] = pd.to_numeric(df_local["Ukenr"], errors="coerce")
                df_local = df_local.dropna(subset=["Ukenr", "tp"])
                if df_local.empty:
                    continue

                latest_idx = df_local["Ukenr"].idxmax()
                latest_row = df_local.loc[latest_idx]

                player_code = table_name.split("_", 1)[1] if "_" in table_name else ws.title
                antall_12 = latest_row.get("Antall 12", None)
                antall_11 = latest_row.get("Antall 11", None)
                antall_10 = latest_row.get("Antall 10", None)
                premie_12 = latest_row.get("12-premie", None)
                premie_11 = latest_row.get("11-premie", None)
                premie_10 = latest_row.get("10-premie", None)
                lev_bonus = latest_row.get("Premie2", latest_row.get("Premie", None))

                output_row: dict[str, object] = {
                    "Navn": str(player_code).upper(),
                    "hp": latest_row.get("hp", None),
                    "up": latest_row.get("up", None),
                    "bp": latest_row.get("bp", None),
                    "straff": latest_row.get("straff", None),
                    "tp": latest_row.get("tp", None),
                    "Tolvere": antall_12,
                    "Elvere": antall_11,
                    "Tiere": antall_10,
                    "12p": premie_12,
                    "11p": premie_11,
                    "10p": premie_10,
                    "12LEV": None,
                    "11LEV": None,
                    "10LEV": None,
                    "LEVp": lev_bonus,
                }
                tp_val = pd.to_numeric(pd.Series([output_row.get("tp")]), errors="coerce").iloc[0]
                levp_val = pd.to_numeric(pd.Series([output_row.get("LEVp")]), errors="coerce").iloc[0]
                output_row["Totalt"] = (
                    (0.0 if pd.isna(tp_val) else float(tp_val))
                    + (0.0 if pd.isna(levp_val) else float(levp_val))
                )
                rows_out.append(output_row)

        if not rows_out:
            return None

        out_df = pd.DataFrame(rows_out)
        for col in [
            "hp",
            "up",
            "bp",
            "straff",
            "tp",
            "Tolvere",
            "Elvere",
            "Tiere",
            "12p",
            "11p",
            "10p",
            "12LEV",
            "11LEV",
            "10LEV",
            "LEVp",
            "Totalt",
        ]:
            if col in out_df.columns:
                out_df[col] = pd.to_numeric(out_df[col], errors="coerce")

        out_df = out_df.sort_values(by="tp", ascending=False).reset_index(drop=True)
        return out_df
    finally:
        wb.close()


def load_weekly_matrix_from_header_range(file_path: str) -> pd.DataFrame | None:
    """Find a weekly matrix by matching the expected header row in workbook cells.

    This targets layouts like:
    Navn, hp, up, bp, straff, tp, Tolvere, Elvere, Tiere, 12p, 11p, 10p,
    12LEV, 11LEV, 10LEV, LEVp, Totalt.
    """
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    try:
        required = ["Navn", "hp", "up", "bp", "straff", "tp"]
        extras = ["Tolvere", "Elvere", "Tiere", "12p", "11p", "10p", "12LEV", "11LEV", "10LEV", "LEVp", "Totalt"]

        for ws in wb.worksheets:
            max_row = min(ws.max_row, 600)
            max_col = min(ws.max_column, 80)

            for r in range(1, max_row + 1):
                row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
                row_text = [str(v).strip() if v is not None else "" for v in row_vals]
                lowered = [t.lower() for t in row_text]

                def find_col(name: str) -> int | None:
                    needle = name.lower()
                    for idx, text in enumerate(lowered):
                        if text == needle:
                            return idx
                    return None

                required_indices = [find_col(col_name) for col_name in required]
                if any(idx is None for idx in required_indices):
                    continue

                has_extra = any(find_col(col_name) is not None for col_name in extras)
                if not has_extra:
                    continue

                all_cols = required + extras
                selected_cols: list[tuple[str, int]] = []
                for col_name in all_cols:
                    idx = find_col(col_name)
                    if idx is not None:
                        selected_cols.append((col_name, idx))

                if not selected_cols:
                    continue

                data_rows: list[list[object]] = []
                for rr in range(r + 1, max_row + 1):
                    row_data = [ws.cell(rr, col_idx + 1).value for _, col_idx in selected_cols]
                    navn_idx = next((i for i, (name, _) in enumerate(selected_cols) if name == "Navn"), None)
                    navn_val = row_data[navn_idx] if navn_idx is not None else None
                    if navn_val is None or str(navn_val).strip() == "":
                        if data_rows:
                            break
                        continue
                    data_rows.append(row_data)

                if not data_rows:
                    continue

                out_df = pd.DataFrame(data_rows, columns=[name for name, _ in selected_cols])
                for numeric_col in [
                    "hp",
                    "up",
                    "bp",
                    "straff",
                    "tp",
                    "Tolvere",
                    "Elvere",
                    "Tiere",
                    "12p",
                    "11p",
                    "10p",
                    "12LEV",
                    "11LEV",
                    "10LEV",
                    "LEVp",
                    "Totalt",
                ]:
                    if numeric_col in out_df.columns:
                        out_df[numeric_col] = pd.to_numeric(out_df[numeric_col], errors="coerce")

                if "Totalt" in out_df.columns:
                    out_df = out_df.sort_values(by="Totalt", ascending=False).reset_index(drop=True)
                elif "tp" in out_df.columns:
                    out_df = out_df.sort_values(by="tp", ascending=False).reset_index(drop=True)

                # Heuristic: reject season-summary tables when looking for weekly table.
                metric_col = "tp" if "tp" in out_df.columns else ("Totalt" if "Totalt" in out_df.columns else None)
                if metric_col is not None:
                    metric_max = pd.to_numeric(out_df[metric_col], errors="coerce").max()
                    if pd.notna(metric_max) and float(metric_max) > 3000.0:
                        continue

                return out_df

        return None
    finally:
        wb.close()


def build_tipper_dashboard_sheet(file_path: str, sheet_name: str = "Dashboard") -> pd.DataFrame:
    """Create or update a dashboard sheet with weekly progression for all tippers.

    The function scans workbook sheets that expose a table named like t_ahh, t_rb,
    etc., reads the columns Dato, hp, up, bp and Totalt, and writes a flattened
    table to a separate Dashboard.xlsm file (not modifying the source).
    """
    source_path = Path(file_path)
    dashboard_path = Path("Dashboard.xlsx")
    
    # Read data from source file (need read_only=False to access tables)
    wb_source = load_workbook(filename=str(source_path), data_only=True, read_only=False)
    
    # Create separate plain workbook for dashboard (no macros = no format issues)
    wb_dashboard = Workbook()
    wb_dashboard.remove(wb_dashboard.active)
    
    try:
        ws_dashboard = wb_dashboard.create_sheet(title=sheet_name)
        ws_dashboard.append(["Tipper", "Dato", "hp", "up", "bp", "Totalt"])

        rows_out: list[dict[str, object]] = []
        for ws in wb_source.worksheets:
            if ws.title == sheet_name:
                continue
            if not ws.tables:
                continue

            for table_name, table in ws.tables.items():
                table_lower = str(table_name).lower()
                if not table_lower.startswith("t_"):
                    continue

                table_ref = table.ref if hasattr(table, "ref") else str(table)
                values = [[cell.value for cell in row] for row in ws[table_ref]]
                if not values:
                    continue

                header, *data = values
                header = [str(c).replace("\xa0", " ").replace("\n", " ").strip() if c is not None else "" for c in header]
                if not data:
                    continue

                df_local = pd.DataFrame(data, columns=header)
                normalized_columns = {
                    c: c for c in df_local.columns
                }
                for candidate in ["Dato", "dato", "Date", "date"]:
                    if candidate in df_local.columns:
                        normalized_columns[candidate] = "Dato"
                for candidate in ["hp", "Hp"]:
                    if candidate in df_local.columns:
                        normalized_columns[candidate] = "hp"
                for candidate in ["up", "Up"]:
                    if candidate in df_local.columns:
                        normalized_columns[candidate] = "up"
                for candidate in ["bp", "Bp"]:
                    if candidate in df_local.columns:
                        normalized_columns[candidate] = "bp"
                for candidate in ["Totalt", "totalt", "Total", "total", "tp", "TP"]:
                    if candidate in df_local.columns:
                        normalized_columns[candidate] = "Totalt"

                df_local = df_local.rename(columns=normalized_columns)

                selected_columns = [col for col in ["Dato", "hp", "up", "bp", "Totalt"] if col in df_local.columns]
                if not selected_columns or "Dato" not in df_local.columns:
                    continue

                player_code = table_name.split("_", 1)[1] if "_" in table_name else ws.title
                df_local = df_local[["Dato", *[col for col in selected_columns if col != "Dato"]]]
                df_local = df_local.dropna(subset=["Dato"], how="all")
                if df_local.empty:
                    continue

                def _coerce_numeric(col: str) -> pd.Series:
                    return pd.to_numeric(df_local[col], errors="coerce")

                for col in ["hp", "up", "bp", "Totalt"]:
                    if col in df_local.columns:
                        df_local[col] = _coerce_numeric(col)

                for _, row in df_local.iterrows():
                    rows_out.append(
                        {
                            "Tipper": str(player_code).upper(),
                            "Dato": row.get("Dato"),
                            "hp": row.get("hp"),
                            "up": row.get("up"),
                            "bp": row.get("bp"),
                            "Totalt": row.get("Totalt"),
                        }
                    )

        if rows_out:
            out_df = pd.DataFrame(rows_out)
            out_df = out_df.sort_values(by=["Tipper", "Dato"], ascending=[True, True]).reset_index(drop=True)
            for row_idx, row in enumerate(out_df.itertuples(index=False), start=2):
                ws_dashboard.append([row.Tipper, row.Dato, row.hp, row.up, row.bp, row.Totalt])
            
            _add_tipper_charts(wb_dashboard, ws_dashboard, out_df)
            
            # Save to separate Dashboard.xlsx file (does not corrupt source)
            wb_dashboard.save(str(dashboard_path))
            return out_df

        # Save empty dashboard
        wb_dashboard.save(str(dashboard_path))
        return pd.DataFrame(columns=["Tipper", "Dato", "hp", "up", "bp", "Totalt"])
    finally:
        wb_source.close()
        wb_dashboard.close()




def _add_tipper_charts(wb_dashboard, ws_dashboard, out_df: pd.DataFrame) -> None:
    """Create a 'Diagrammer' sheet with one matplotlib line chart image per tipper."""
    import io
    from openpyxl.drawing.image import Image as XLImage

    ws_charts = wb_dashboard.create_sheet(title="Diagrammer")
    tippers = sorted(out_df["Tipper"].unique())

    img_width_px = 480
    img_height_px = 300
    # Each image is ~480px wide. At ~7px per Excel col unit and col width ~8.43: ~9 cols wide.
    # At ~15px per Excel row: ~20 rows tall.
    col_offsets = [1, 10]   # Column A and J
    row_step = 20           # Rows per chart slot

    for tidx, tipper in enumerate(tippers):
        tipper_df = out_df[out_df["Tipper"] == tipper].sort_values("Dato").reset_index(drop=True)
        totalt = pd.to_numeric(tipper_df["Totalt"], errors="coerce")
        if totalt.dropna().empty:
            continue

        # Format x-axis labels
        labels = [
            d.strftime("%d.%m") if hasattr(d, "strftime") else str(d)
            for d in tipper_df["Dato"]
        ]

        fig, ax = plt.subplots(figsize=(img_width_px / 96, img_height_px / 96))
        ax.plot(range(len(totalt)), totalt.values, marker="o", markersize=4,
                linewidth=2, color="#2563EB")
        ax.set_title(tipper, fontsize=13, fontweight="bold", pad=8)
        ax.set_ylabel("Totalt poeng", fontsize=9)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7)
        ax.yaxis.grid(True, linestyle="--", alpha=0.5)
        ax.set_axisbelow(True)
        fig.tight_layout()

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=96)
        plt.close(fig)
        buf.seek(0)

        img = XLImage(buf)
        img.width = img_width_px
        img.height = img_height_px

        col_idx = tidx % 2
        row_idx = tidx // 2
        # Convert col index to letter (A=1, J=10)
        col_letter = chr(ord("A") + col_offsets[col_idx] - 1)
        row_anchor = row_idx * row_step + 1
        ws_charts.add_image(img, f"{col_letter}{row_anchor}")


def _add_dashboard_charts(wb, ws_dashboard, df: pd.DataFrame) -> None:
    """Add visualization charts to the Dashboard sheet.
    
    Creates line charts showing tipper progression over time.
    """
    if df.empty or ws_dashboard.max_row < 3:
        return
    
    try:
        data_start_row = 2
        data_end_row = ws_dashboard.max_row
        
        # Chart 1: Totalt points over time for all tippers
        chart_totalt = LineChart()
        chart_totalt.title = "Sesongprogresjon - Totalt Poeng"
        chart_totalt.style = 12
        chart_totalt.y_axis.title = "Poeng"
        chart_totalt.x_axis.title = "Dato"
        chart_totalt.height = 12
        chart_totalt.width = 20
        
        # Get unique tippers
        tippers = sorted(df["Tipper"].unique())
        
        col_date = 2  # Dato column
        col_totalt = 6  # Totalt column
        
        for tipper in tippers:
            tipper_rows = df[df["Tipper"] == tipper].index.tolist()
            if not tipper_rows:
                continue
            
            first_row = min(tipper_rows) + 2
            last_row = max(tipper_rows) + 2
            
            values = Reference(ws_dashboard, min_col=col_totalt, min_row=first_row, max_row=last_row)
            series = chart_totalt.add_data(values, titles_from_data=False)
            if series:
                series.title = tipper
        
        chart_totalt.set_categories(Reference(ws_dashboard, min_col=col_date, min_row=data_start_row, max_row=data_end_row))
        ws_dashboard.add_chart(chart_totalt, "H2")
        
        # Chart 2: HP/UP/BP breakdown for most recent week
        chart_breakdown = BarChart()
        chart_breakdown.type = "col"
        chart_breakdown.title = "Siste Uke - HP/UP/BP Fordeling"
        chart_breakdown.style = 10
        chart_breakdown.y_axis.title = "Poeng"
        chart_breakdown.x_axis.title = "Tipper"
        chart_breakdown.height = 10
        chart_breakdown.width = 20
        
        latest_df = df.drop_duplicates(subset=["Tipper"], keep="last").sort_values("Tipper")
        if not latest_df.empty:
            chart_start_row = data_end_row + 3
            chart_data_start = chart_start_row + 1
            
            ws_dashboard.cell(chart_start_row, 1).value = "Tipper"
            ws_dashboard.cell(chart_start_row, 2).value = "HP"
            ws_dashboard.cell(chart_start_row, 3).value = "UP"
            ws_dashboard.cell(chart_start_row, 4).value = "BP"
            
            for row_offset, (_, row) in enumerate(latest_df.iterrows()):
                ws_dashboard.cell(chart_data_start + row_offset, 1).value = row["Tipper"]
                ws_dashboard.cell(chart_data_start + row_offset, 2).value = row.get("hp", 0) or 0
                ws_dashboard.cell(chart_data_start + row_offset, 3).value = row.get("up", 0) or 0
                ws_dashboard.cell(chart_data_start + row_offset, 4).value = row.get("bp", 0) or 0
            
            last_chart_row = chart_data_start + len(latest_df) - 1
            categories = Reference(ws_dashboard, min_col=1, min_row=chart_data_start, max_row=last_chart_row)
            hp_data = Reference(ws_dashboard, min_col=2, min_row=chart_start_row, max_row=last_chart_row)
            up_data = Reference(ws_dashboard, min_col=3, min_row=chart_start_row, max_row=last_chart_row)
            bp_data = Reference(ws_dashboard, min_col=4, min_row=chart_start_row, max_row=last_chart_row)
            
            chart_breakdown.add_data(hp_data, titles_from_data=True)
            chart_breakdown.add_data(up_data, titles_from_data=True)
            chart_breakdown.add_data(bp_data, titles_from_data=True)
            chart_breakdown.set_categories(categories)
            chart_breakdown.grouping = "stacked"
            
            ws_dashboard.add_chart(chart_breakdown, "H30")
    except Exception as e:
        logger.warning(f"Advarsel: Klarte ikke å legge til dashboard-diagrammer: {e}")



def prepare_report_df(input_df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    df_local = input_df.copy()
    df_local.columns = [
        str(c).replace("\xa0", " ").replace("\n", " ").strip() if c is not None else ""
        for c in df_local.columns
    ]

    non_empty_headers = [c for c in df_local.columns if c and c.lower() != "nan"]
    unnamed_count = sum(1 for c in df_local.columns if c.lower().startswith("unnamed:"))
    if (len(non_empty_headers) <= 2 or unnamed_count >= max(1, len(df_local.columns) // 2)) and len(df_local) > 0:
        promoted = [
            str(c).replace("\xa0", " ").replace("\n", " ").strip()
            if c is not None
            else ""
            for c in df_local.iloc[0].tolist()
        ]
        if any(promoted):
            df_local.columns = promoted
            df_local = df_local.iloc[1:].reset_index(drop=True)

    df_local = df_local.rename(columns={df_local.columns[0]: "Navn"})
    df_local["Navn"] = df_local["Navn"].astype(str).str.strip()
    df_local = df_local[
        ~df_local["Navn"].str.lower().isin({"", "none", "nan", "null"})
    ]

    total_col_local = None
    for c in df_local.columns:
        lc = c.lower().strip()
        if lc == "totalt" or "totalt" in lc:
            total_col_local = c
            break

    if total_col_local is None:
        raise KeyError(f"Fant ikke 'Totalt'-kolonne. Kolonner i datasettet: {list(df_local.columns)}")

    df_local[total_col_local] = pd.to_numeric(df_local[total_col_local], errors="coerce")
    df_local = df_local.dropna(subset=[total_col_local])

    df_local = (
        df_local.sort_values(by=total_col_local, ascending=False)
        .drop_duplicates(subset=["Navn"], keep="first")
        .sort_values(by=total_col_local, ascending=False)
        .reset_index(drop=True)
    )

    return df_local, total_col_local


def render_section(
    df_local: pd.DataFrame,
    total_col_local: str,
    title: str,
    y_max: float = 25000.0,
) -> tuple[str, str]:
    fig, ax = plt.subplots(figsize=(7, 4))

    bar_colors = ["#d4af37", "#c0c0c0", "#cd7f32"] + ["#175c5f"] * max(0, len(df_local) - 3)
    bars = ax.bar(df_local["Navn"], df_local[total_col_local], color=bar_colors[: len(df_local)], width=0.6)

    legend_handles = []
    if len(df_local) >= 1: 
        legend_handles.append(Patch(facecolor="#d4af37", label="1. plass"))
    if len(df_local) >= 2:
        legend_handles.append(Patch(facecolor="#c0c0c0", label="2. plass"))
    if len(df_local) >= 3:
        legend_handles.append(Patch(facecolor="#cd7f32", label="3. plass"))
    if legend_handles:
        ax.legend(handles=legend_handles, loc="upper right", frameon=True)

    ax.set_title(title)
    ax.set_xlabel(df_local.columns[0])
    ax.set_ylabel(total_col_local)

    y_min = 0.0
    ax.set_ylim(y_min, y_max)

    label_offset = max((y_max - y_min) * 0.01, 1.0)
    for bar, value in zip(bars, df_local[total_col_local]):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            float(value) + label_offset,
            f"{int(round(float(value)))}",
            ha="center",
            va="bottom",
            fontsize=7,
            color="#222222",
        )

    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format="png")
    img_buffer.seek(0)
    img_base64_local = base64.b64encode(img_buffer.getvalue()).decode("utf-8")
    plt.close(fig)

    display_df_local = df_local.copy()
    numeric_view = display_df_local.apply(lambda col: pd.to_numeric(col, errors="coerce"))
    display_df_local = display_df_local.mask(numeric_view.eq(0))
    display_df_local = display_df_local.replace({None: ""}).fillna("")
    display_df_local = display_df_local.replace(
        to_replace=r"^\s*(none|null|nan)\s*$",
        value="",
        regex=True,
    )
    display_df_local = display_df_local.replace(to_replace=r"^\s*0(?:\.0+)?\s*$", value="", regex=True)

    html_table_local = display_df_local.to_html(
        classes="styled-table",
        index=False,
        na_rep="",
        float_format="{:.0f}".format,
    )

    section_html_local = f"""
    <div class="chart-container">
        <img src="data:image/png;base64,{img_base64_local}" alt="{title}">
    </div>
    <h2>{title}</h2>
    <div class="table-wrap">
        {html_table_local}
    </div>
    """

    return section_html_local, img_base64_local


def build_weekly_section_html(source_workbook: str | None, report_week: int) -> tuple[str, pd.DataFrame | None, str | None]:
    """Build the weekly section using a prioritized fallback chain."""
    if source_workbook is None:
        return (
            "<p style='color:#555; text-align:left;'>"
            "Fant ingen kildefil for navngitt område 'Sist_uke'."
            "</p>",
            None,
            None,
        )

    weekly_error = None
    weekly_df = None
    weekly_total_col = None

    # 1) Preferred source: named ranges that should contain this week's full table.
    for candidate in WEEKLY_CANDIDATES:
        try:
            weekly_raw_df = load_named_range_to_df(
                source_workbook,
                candidate,
                allow_generic_totalt_fallback=False,
            )
            weekly_df, weekly_total_col = prepare_report_df(weekly_raw_df)
            if len(weekly_df) > 0:
                logger.info(f"Info: Bruker ukesdata fra navngitt område '{candidate}'.")
                break
        except Exception as exc:
            weekly_error = exc

    if weekly_df is not None and weekly_total_col is not None and len(weekly_df) > 0:
        weekly_section_html, _ = render_section(
            weekly_df,
            weekly_total_col,
            f"Sist uke - Uke {report_week} - Tippernes resultater",
            y_max=2000.0,
        )
        return weekly_section_html, weekly_df, weekly_total_col

    # 2) Secondary source: legacy workbook table used in some files.
    try:
        weekly_raw_df = load_table_from_workbook(source_workbook, "Summasummarium5")
        weekly_df, weekly_total_col = prepare_report_df(weekly_raw_df)
        weekly_section_html, _ = render_section(
            weekly_df,
            weekly_total_col,
            f"Sist uke - Uke {report_week} - Tippernes resultater",
            y_max=2000.0,
        )
        logger.info("Info: Bruker ukesdata fra tabell 'Summasummarium5'.")
        return weekly_section_html, weekly_df, weekly_total_col
    except Exception as table_exc:
        # 3) Preferred fallback: look for the exact weekly matrix layout in workbook cells.
        weekly_matrix_df = load_weekly_matrix_from_header_range(source_workbook)
        if weekly_matrix_df is None and Path(WORKBOOK_FILE).exists():
            weekly_matrix_df = load_weekly_matrix_from_header_range(WORKBOOK_FILE)

        if weekly_matrix_df is not None and len(weekly_matrix_df) > 0:
            metric_col = "Totalt" if "Totalt" in weekly_matrix_df.columns else "tp"
            weekly_section_html, _ = render_section(
                weekly_matrix_df,
                metric_col,
                f"Sist uke - Uke {report_week} - Tippernes resultater",
                y_max=max(float(pd.to_numeric(weekly_matrix_df[metric_col], errors='coerce').max()) + 200.0, 1000.0),
            )
            logger.info("Info: Bruker fallback-data fra ukematrise med Navn/hp/up/bp/straff/tp-kolonner.")
            return weekly_section_html, weekly_matrix_df, metric_col

        # 4) Fallback: reconstruct this week's table from player history tables.
        player_weekly_df = load_latest_weekly_from_player_tables(source_workbook)
        if player_weekly_df is not None and len(player_weekly_df) > 0:
            weekly_section_html, _ = render_section(
                player_weekly_df,
                "tp",
                f"Sist uke - Tippernes resultater - Uke {report_week}",
                y_max=max(float(player_weekly_df["tp"].max()) + 200.0, 1000.0),
            )
            logger.info("Info: Bruker fallback-data fra spillerhistorikk-tabeller (f_*/t_*).")
            return weekly_section_html, player_weekly_df, "tp"

        # 5) Last fallback: winner counts only (reduced column set).
        winner_counts_df = load_weekly_winner_counts_from_named_range(source_workbook, "Ukevinnere")
        if winner_counts_df is not None and len(winner_counts_df) > 0:
            weekly_section_html, _ = render_section(
                winner_counts_df,
                "Totalt",
                f"Ukevinnere hittil - Uke {report_week}",
                y_max=max(float(winner_counts_df["Totalt"].max()) + 1.0, 5.0),
            )
            logger.info("Info: Bruker fallback-data fra navngitt område 'Ukevinnere'.")
            return weekly_section_html, winner_counts_df, "Totalt"

        if weekly_error is not None:
            logger.warning(f"Advarsel: Klarte ikke å bygge seksjon for 'Sist_uke': {weekly_error}")
        logger.warning(f"Advarsel: Klarte heller ikke å bruke tabell 'Summasummarium5': {table_exc}")
        return (
            "<p style='color:#555; text-align:left;'>"
            "Fant ikke gyldige data for navngitt område 'Sist_uke'."
            "</p>",
            None,
            None,
        )


def build_player_history_entry(
    report_week: int,
    report_year: int,
    weekly_df: pd.DataFrame | None,
    total_col: str | None,
) -> dict[str, object]:
    if weekly_df is None or total_col is None or weekly_df.empty:
        return {"week": report_week, "year": report_year, "rows": []}

    rows: list[dict[str, object]] = []
    for _, row in weekly_df.iterrows():
        name = str(row.get("Navn", "")).strip()
        if not name or name.lower() in {"", "none", "nan", "null"}:
            continue
        total_value = pd.to_numeric(row.get(total_col), errors="coerce")
        rows.append(
            {
                "Navn": name,
                "Totalt": float(total_value) if pd.notna(total_value) else 0.0,
                "Plass": 0,
            }
        )

    if not rows:
        return {"week": report_week, "year": report_year, "rows": []}

    rows.sort(key=lambda item: item["Totalt"], reverse=True)
    for idx, row in enumerate(rows, start=1):
        row["Plass"] = idx

    return {"week": report_week, "year": report_year, "rows": rows}


def normalize_player_name(player_name: str) -> str:
    return re.sub(r"\s+", " ", str(player_name).strip().upper())


def load_player_history(history_path: Path) -> dict[str, object]:
    if not history_path.exists():
        return {"players": {}}

    try:
        with history_path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
            if isinstance(data, dict):
                return data
    except Exception:
        logger.warning(f"Advarsel: Klarte ikke å lese spillerhistorikk fra '{history_path}'.")

    return {"players": {}}


def update_player_history_archive(
    history_path: Path,
    report_week: int,
    report_year: int,
    weekly_df: pd.DataFrame | None,
    total_col: str | None,
) -> dict[str, object]:
    history = load_player_history(history_path)
    players = history.setdefault("players", {})
    if not isinstance(players, dict):
        players = {}
        history["players"] = players

    entry = build_player_history_entry(report_week, report_year, weekly_df, total_col)
    if not entry.get("rows"):
        return history

    for row in entry["rows"]:
        player_name = normalize_player_name(str(row.get("Navn", "")))
        if not player_name:
            continue

        player_records = players.setdefault(player_name, [])
        if not isinstance(player_records, list):
            player_records = []
            players[player_name] = player_records

        already_exists = any(
            record.get("week") == report_week and record.get("year") == report_year
            for record in player_records
        )
        if already_exists:
            continue

        player_records.append(
            {
                "week": report_week,
                "year": report_year,
                "plass": row.get("Plass"),
                "totalt": row.get("Totalt"),
                "navn": str(row.get("Navn", "")),
            }
        )

    for player_name, records in players.items():
        if isinstance(records, list):
            records.sort(key=lambda item: (item.get("year", 0), item.get("week", 0)))

    history["players"] = {
        player_name: players[player_name]
        for player_name in sorted(players.keys(), key=lambda item: str(item))
    }

    history_path.parent.mkdir(parents=True, exist_ok=True)
    with history_path.open("w", encoding="utf-8") as handle:
        json.dump(history, handle, ensure_ascii=False, indent=2)

    return history


def build_player_history_section_html(history_path: Path) -> str:
    history = load_player_history(history_path)
    players = history.get("players", {})
    if not isinstance(players, dict) or not players:
        return "<p style='color:#555; text-align:left;'>Ingen spillerhistorikk er tilgjengelig ennå.</p>"

    options = []
    for player_name in sorted(players.keys(), key=lambda item: str(item)):
        safe_name = html.escape(str(player_name))
        options.append(f'<option value="{safe_name}">{safe_name}</option>')

    player_data_json = json.dumps(players, ensure_ascii=False)
    options_html = "".join(options)

    return """
    <div class="player-history-panel" style="margin-top: 24px; padding-top: 16px; border-top: 1px solid #d7d7d7;">
        <h2>Velg tipper og se historikk</h2>
        <label for="player-history-select" style="font-weight: 600; display: block; margin-bottom: 8px;">Tipper</label>
        <select id="player-history-select" style="padding: 8px 10px; min-width: 240px; border: 1px solid #175c5f; border-radius: 6px;">
            %s
        </select>
        <div id="player-history-output" style="margin-top: 14px;"></div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
    <script>
        const playerHistoryData = %s;
        const historySelect = document.getElementById('player-history-select');
        const historyOutput = document.getElementById('player-history-output');

        function escapeHtml(value) {
            return String(value)
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;');
        }

        function renderPlayerHistory(playerName) {
            const rows = playerHistoryData[playerName] || [];
            if (!rows.length) {
                historyOutput.innerHTML = '<p style=\"color:#555;\">Ingen historikk funnet for denne tipperen.</p>';
                return;
            }

            const sortedRows = [...rows].sort((a, b) => (a.week ?? 0) - (b.week ?? 0));
            const tableRows = sortedRows.map((row) => {
                const weekLabel = row.week ?? '-';
                const placeLabel = row.plass ?? '-';
                const totalLabel = Number(row.totalt ?? 0).toLocaleString('nb-NO');
                return `<tr><td>${escapeHtml(weekLabel)}</td><td>${escapeHtml(placeLabel)}</td><td>${escapeHtml(totalLabel)}</td></tr>`;
            }).join('');

            const weeks = sortedRows.map((row) => `Uke ${row.week ?? '-'}`);
            const ranks = sortedRows.map((row) => Number(row.plass ?? 0));
            const totals = sortedRows.map((row) => Number(row.totalt ?? 0));

            historyOutput.innerHTML = `
                <h3 style=\"margin-bottom: 8px;\">${escapeHtml(playerName)}</h3>
                <div style=\"display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 16px; margin-top: 12px;\">
                    <div style=\"background: #f8f8f8; padding: 10px; border-radius: 8px;\">
                        <canvas id=\"rank-chart\" height=\"220\"></canvas>
                    </div>
                    <div style=\"background: #f8f8f8; padding: 10px; border-radius: 8px;\">
                        <canvas id=\"points-chart\" height=\"220\"></canvas>
                    </div>
                </div>
                <div class=\"table-wrap\" style=\"margin-top: 14px;\">
                    <table class=\"styled-table\" style=\"margin-top: 0;\">
                        <thead><tr><th>Uke</th><th>Plass</th><th>Totalt</th></tr></thead>
                        <tbody>${tableRows}</tbody>
                    </table>
                </div>
            `;

            setTimeout(() => {
                const rankCtx = document.getElementById('rank-chart');
                const pointsCtx = document.getElementById('points-chart');

                if (rankCtx) {
                    new Chart(rankCtx, {
                        type: 'line',
                        data: {
                            labels: weeks,
                            datasets: [{
                                label: 'Plassering',
                                data: ranks,
                                borderColor: '#175c5f',
                                backgroundColor: 'rgba(23, 92, 95, 0.2)',
                                fill: true,
                                tension: 0.25
                            }]
                        },
                        options: {
                            responsive: true,
                            plugins: { legend: { display: false } },
                            scales: {
                                y: {
                                    reverse: true,
                                    beginAtZero: true,
                                    ticks: { stepSize: 1 }
                                }
                            }
                        }
                    });
                }

                if (pointsCtx) {
                    new Chart(pointsCtx, {
                        type: 'bar',
                        data: {
                            labels: weeks,
                            datasets: [{
                                label: 'Poeng',
                                data: totals,
                                backgroundColor: '#d4af37',
                                borderColor: '#b8910f',
                                borderWidth: 1
                            }]
                        },
                        options: {
                            responsive: true,
                            plugins: { legend: { display: false } },
                            scales: {
                                y: { beginAtZero: true }
                            }
                        }
                    });
                }
            }, 0);
        }

        if (historySelect && historyOutput) {
            historySelect.addEventListener('change', (event) => renderPlayerHistory(event.target.value));
            renderPlayerHistory(historySelect.value || historySelect.options[0]?.value);
        }
    </script>
    """ % (options_html, player_data_json)


def export_html_to_pdf(html_path: str, pdf_path: str) -> None:
    try:
        from weasyprint import HTML  # type: ignore

        HTML(filename=html_path).write_pdf(pdf_path)
        logger.info(f"PDF eksportert: '{pdf_path}' (weasyprint).")
        return
    except Exception:
        pass

    wkhtmltopdf_path = resolve_wkhtmltopdf_path()
    if wkhtmltopdf_path:
        try:
            subprocess.run(
                [
                    wkhtmltopdf_path,
                    "--enable-local-file-access",
                    "--print-media-type",
                    "--orientation",
                    "Portrait",
                    html_path,
                    pdf_path,
                ],
                check=True,
            )
            logger.info(f"PDF eksportert: '{pdf_path}' (wkhtmltopdf).")
            return
        except Exception as exc:
            logger.warning(f"Advarsel: PDF-eksport feilet via wkhtmltopdf: {exc}")
            return

    logger.info(
        "Info: PDF ikke laget automatisk. Installer 'weasyprint' eller 'wkhtmltopdf' "
        "for PDF-eksport."
    )


def resolve_wkhtmltopdf_path() -> str | None:
    candidates: list[Path] = []

    env_path = os.environ.get("WKHTMLTOPDF_PATH")
    if env_path:
        candidates.append(Path(env_path))

    which_path = shutil.which("wkhtmltopdf")
    if which_path:
        candidates.append(Path(which_path))

    candidates.extend(
        [
            Path(r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"),
            Path(r"C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe"),
        ]
    )

    for candidate in candidates:
        if candidate.is_file():
            return str(candidate)

    return None


def pdf_export_available() -> bool:
    return importlib.util.find_spec("weasyprint") is not None or resolve_wkhtmltopdf_path() is not None


file_exists = Path(file).exists()
should_refresh = refresh_on_start or not file_exists

# --- 1) Load and refresh input data from workbook ---
if should_refresh:
    if not file_exists:
        logger.info("Info: stps_tolk.xlsx mangler. Prøver å opprette den fra 2027-kilde.")
        startup_df = update_stps_tolk_on_start(file, sheet_name)
        if startup_df is not None:
            df = startup_df.copy()
        else:
            if not Path(file).exists():
                raise FileNotFoundError(
                    "stps_tolk.xlsx finnes ikke, og automatisk oppretting fra 2027-kilde feilet. "
                    "Kjør med en gyldig kildefil tilgjengelig (f.eks. STPS_2027.xlsm)."
                )
            df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
    else:
        source_workbook_for_refresh = find_source_workbook()
        if source_workbook_for_refresh is not None:
            refresh_linked_workbooks_via_excel(source_workbook_for_refresh, file)
        else:
            logger.info("Info: Fant ingen kildefil å åpne før stps_tolk.xlsx for link-oppdatering.")

        logger.info("Info: Leser data direkte fra stps_tolk.xlsx (ingen manuell omskriving av arket).")
        df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
else:
    logger.info("Info: Leser direkte fra stps_tolk.xlsx. Bruk --refresh for å hente fra 2027-kilde.")
    df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")

# --- 2) Prepare runtime metadata and output filenames ---
now = datetime.now()
report_updated_at = now.strftime("%d.%m.%Y %H:%M:%S")
report_week = now.isocalendar().week
report_year = now.year
pdf_ready_hint = pdf_export_available()
latest_report_html_filename = "report_2027.html"
latest_report_pdf_filename = "report_2027.pdf"
report_html_filename = build_weekly_filename("report", ".html", report_week, report_year)
report_pdf_filename = build_weekly_filename("report", ".pdf", report_week, report_year)

# --- 3) Build summary section (chart + table) ---
summary_df, summary_total_col = prepare_report_df(df)
summary_section_html, _ = render_section(
    summary_df,
    summary_total_col,
    f"Sammendrag - Tippernes resultater 2026/2027 - Uke {report_week}",
)

# --- 4) Build weekly section via fallback chain ---
source_workbook = find_source_workbook()
weekly_section_html, weekly_df, weekly_total_col = build_weekly_section_html(source_workbook, report_week)
player_history_archive = update_player_history_archive(
    PLAYER_HISTORY_FILE,
    report_week,
    report_year,
    weekly_df,
    weekly_total_col,
)
player_history_section_html = build_player_history_section_html(PLAYER_HISTORY_FILE)

# --- 5) Render complete report HTML from template ---
html_content = render_report_html(
    REPORT_TEMPLATE_PATH,
    build_logo_src(resolve_logo_path()),
    report_week,
    latest_report_pdf_filename,
    pdf_ready_hint,
    report_updated_at,
    summary_section_html,
    weekly_section_html,
    player_history_section_html,
)

# --- 6) Write HTML outputs and export PDF ---
with open(report_html_filename, "w", encoding="utf-8") as f:
    f.write(html_content)

with open(latest_report_html_filename, "w", encoding="utf-8") as f:
    f.write(html_content)

export_html_to_pdf(report_html_filename, report_pdf_filename)
shutil.copy2(report_pdf_filename, latest_report_pdf_filename)

# --- 7) Save weekly workbook snapshot when source is xlsm ---
if source_workbook is not None and source_workbook.lower().endswith(".xlsm"):
    workbook_copy = create_weekly_workbook_copy(source_workbook, report_week, report_year)
    if workbook_copy is not None:
        logger.info(f"Kopierte kildearbeidsbok til '{workbook_copy}'.")

logger.info(f"HTML report successfully generated as '{report_html_filename}'!")
