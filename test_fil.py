
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("SkorgenTippelag.xlsm", data_only=True)
ws = wb["Statistikk"]

table = ws.tables["t_tp"]

data = ws[table.ref]

df = pd.DataFrame([[cell.value for cell in row] for row in data])

df.columns = df.iloc[0]
df = df[1:]