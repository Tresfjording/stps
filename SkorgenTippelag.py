import pandas as pd
import matplotlib.pyplot as plt
import datetime
from openpyxl import load_workbook

# Parse repeated horizontal blocks (kuponger)

FILE = "stps_tolk.xlsx"
SHEET = "kuponger"

# Konfig: vertikal og horisontal offset
START_ROW = 1            # A1-basis
WEEK_STEP = 23           # rader per uke (vertikal forskyvning)
BLOCKS = 9               # antall horisontale blokker
BLOCK_START_COL = 14     # N = kolonne 14 (1-basert)
BLOCK_SPACING = 10       # start->start mellom blokker
BLOCK_WIDTH = 8          # bredde på hver blokk
ROWS_PER_BLOCK = 12      # rader i hver blokk (f.eks. 5..16)


def read_blocks(file=FILE, sheet=SHEET):
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active

    weeks = []

    for week_idx in range(100):
        row_start = START_ROW + week_idx * WEEK_STEP
        r0 = row_start + 4
        r1 = r0 + ROWS_PER_BLOCK - 1

        # stopp hvis ukens område er tomt i kolonne A (ingen data)
        colA_vals = [ws.cell(row=r, column=1).value for r in range(r0, r1+1)]
        if all(v in (None, "") for v in colA_vals):
            break

        week = {"week_index": week_idx, "blocks": []}

        for block_idx in range(BLOCKS):
            start_col = BLOCK_START_COL + block_idx * BLOCK_SPACING
            block_vals = []
            for r in range(r0, r1+1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(start_col, start_col + BLOCK_WIDTH)]
                block_vals.append(row_vals)

            # sjekk om blokken inneholder noe
            has_data = any(any(cell not in (None, "") for cell in row) for row in block_vals)
            week["blocks"].append({
                "block_idx": block_idx,
                "start_col": start_col,
                "has_data": has_data,
                "values": block_vals
            })

        weeks.append(week)

    return weeks


def print_summary(weeks):
    for w in weeks:
        print(f"Week {w['week_index']}:")
        for b in w["blocks"]:
            print(f"  Block {b['block_idx']} start_col={b['start_col']} has_data={b['has_data']}")


if __name__ == '__main__':
    weeks = read_blocks()
    print(f"Found {len(weeks)} weeks with data")
    print_summary(weeks)

    # --- Lag en kompakt årsoversikt og skriv til Excel (tippelag.xlsx)
    import os

    summary_rows = []
    for w in weeks:
        for b in w['blocks']:
            non_empty = sum(1 for row in b['values'] for cell in row if cell not in (None, ""))
            summary_rows.append({
                'week': w['week_index'],
                'block': b['block_idx'],
                'start_col': b['start_col'],
                'has_data': b['has_data'],
                'non_empty_cells': non_empty
            })

    df_summary = pd.DataFrame(summary_rows)

    out_file = 'tippelag.xlsx'
    try:
        if os.path.exists(out_file):
            with pd.ExcelWriter(out_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_summary.to_excel(writer, sheet_name='Skorgen_Aarsoversikt', index=False)
        else:
            with pd.ExcelWriter(out_file, engine='openpyxl', mode='w') as writer:
                df_summary.to_excel(writer, sheet_name='Skorgen_Aarsoversikt', index=False)

        print(f"Wrote summary to {out_file} sheet Skorgen_Aarsoversikt")
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_file = f'tippelag_Skorgen_Aarsoversikt_{timestamp}.xlsx'
        df_summary.to_excel(alt_file, sheet_name='Skorgen_Aarsoversikt', index=False)
        print(f"Could not write to {out_file} (file may be open). Wrote to {alt_file} instead.")


